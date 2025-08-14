from flask import Flask, request, jsonify
from flask_cors import CORS
import uuid
import datetime
import base64
import json
import os
import threading
from openpyxl import Workbook, load_workbook
import fcntl
from werkzeug.utils import secure_filename

app = Flask(__name__)
CORS(app)  # 启用跨域支持

# 确保logs目录存在
if not os.path.exists('logs'):
    os.makedirs('logs')

LOGS_DIR = 'logs'
EXCEL_FILE = os.path.join(LOGS_DIR, 'training_records.xlsx')
LOCK_FILE = os.path.join(LOGS_DIR, 'training_records.lock')
UPLOADS_DIR = os.path.join(LOGS_DIR, 'uploads')
os.makedirs(UPLOADS_DIR, exist_ok=True)

excel_lock = threading.Lock()

EXCEL_SHEETS = {
    'starts': ['write_time', 'timestamp', 'session_id', 'employee_id'],
    'steps': ['write_time', 'timestamp', 'session_id', 'employee_id', 'step_number', 'step_name'],
    'forms': ['write_time', 'timestamp', 'session_id', 'employee_id', 'step_number', 'field_name', 'field_value'],
    'uploads': ['write_time', 'timestamp', 'session_id', 'employee_id', 'step_number', 'field_name', 'original_filename', 'saved_path', 'file_size_bytes', 'mime_type'],
    'completions': ['write_time', 'timestamp', 'session_id', 'employee_id', 'verification_code'],
    'closures': ['write_time', 'timestamp', 'session_id', 'employee_id', 'step_number'],
    'operations': ['write_time', 'timestamp', 'session_id', 'employee_id', 'operation_type', 'step_number', 'name', 'value', 'extra']
}


def ensure_excel_workbook():
    """Ensure the Excel workbook exists with required sheets and headers."""
    os.makedirs(LOGS_DIR, exist_ok=True)
    # Create lock file if not exists
    with open(LOCK_FILE, 'a'):
        pass
    with excel_lock:
        with open(LOCK_FILE, 'r+') as lf:
            fcntl.flock(lf, fcntl.LOCK_EX)
            try:
                if not os.path.exists(EXCEL_FILE):
                    wb = Workbook()
                    # Remove default sheet
                    default_sheet = wb.active
                    wb.remove(default_sheet)
                    # Create sheets with headers
                    for sheet_name, headers in EXCEL_SHEETS.items():
                        ws = wb.create_sheet(title=sheet_name)
                        ws.append(headers)
                    wb.save(EXCEL_FILE)
                else:
                    # Ensure all sheets and headers exist
                    wb = load_workbook(EXCEL_FILE)
                    for sheet_name, headers in EXCEL_SHEETS.items():
                        if sheet_name not in wb.sheetnames:
                            ws = wb.create_sheet(title=sheet_name)
                            ws.append(headers)
                        else:
                            ws = wb[sheet_name]
                            # If sheet is empty or header missing, write headers
                            if ws.max_row == 0 or (ws.max_row == 1 and all(cell.value is None for cell in ws[1])):
                                ws.append(headers)
                    wb.save(EXCEL_FILE)
            finally:
                fcntl.flock(lf, fcntl.LOCK_UN)


def append_to_excel(sheet_name: str, record: dict):
    """Append a record to the given Excel sheet in a concurrency-safe way."""
    ensure_excel_workbook()
    with excel_lock:
        with open(LOCK_FILE, 'r+') as lf:
            fcntl.flock(lf, fcntl.LOCK_EX)
            try:
                wb = load_workbook(EXCEL_FILE)
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(title=sheet_name)
                    ws.append(EXCEL_SHEETS[sheet_name])
                else:
                    ws = wb[sheet_name]
                # Ensure header exists
                if ws.max_row == 0:
                    ws.append(EXCEL_SHEETS[sheet_name])
                row = [record.get(col) for col in EXCEL_SHEETS[sheet_name]]
                ws.append(row)
                wb.save(EXCEL_FILE)
            finally:
                fcntl.flock(lf, fcntl.LOCK_UN)


def log_operation(operation_type: str, session_id: str, employee_id: str, step_number: int, name: str = '', value: str = '', extra: str = '', timestamp: str = None):
    """Log an aggregate operation row into operations sheet."""
    ts = timestamp or datetime.datetime.now().isoformat()
    append_to_excel('operations', {
        'write_time': datetime.datetime.now().isoformat(),
        'timestamp': ts,
        'session_id': session_id,
        'employee_id': employee_id,
        'operation_type': operation_type,
        'step_number': step_number,
        'name': name,
        'value': value,
        'extra': extra
    })

# 模拟数据库存储
training_data = {
    'sessions': {},
    'steps': [],
    'forms': [],
    'uploads': [],
    'completions': [],
    'closures': []
}

# 工具函数：验证API Key
def verify_api_key():
    api_key = request.headers.get('X-API-Key')
    # 在实际应用中，应该验证API Key的有效性
    # 这里简化处理，只要存在即可
    return api_key is not None

# 工具函数：验证请求数据
def validate_request_data(required_fields):
    data = request.get_json()
    if not data:
        return False, "缺少JSON数据"
    
    for field in required_fields:
        if field not in data:
            return False, f"缺少必需字段: {field}"
    
    return True, data

# 工具函数：将记录写入文件
def write_record_to_file(record_type, data):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d")
    filename = f"logs/training_{record_type}_{timestamp}.log"
    
    # 添加写入时间
    data['write_time'] = datetime.datetime.now().isoformat()
    
    with open(filename, 'a', encoding='utf-8') as f:
        f.write(json.dumps(data, ensure_ascii=False) + '\n')

# API 1: 记录用户开始演练
@app.route('/api/training/start', methods=['POST'])
def record_start():
    if not verify_api_key():
        return jsonify({'status': 'error', 'message': '无效的API Key'}), 401
    
    is_valid, data = validate_request_data(['employee_id'])
    if not is_valid:
        return jsonify({'status': 'error', 'message': data}), 400
    
    # 创建新的会话ID
    session_id = str(uuid.uuid4())
    
    # 存储会话信息
    training_data['sessions'][session_id] = {
        'employee_id': data['employee_id'],
        'start_time': data.get('timestamp', datetime.datetime.now().isoformat())
    }
    
    # 写入文件
    start_record = {
        'session_id': session_id,
        'employee_id': data['employee_id'],
        'timestamp': data.get('timestamp', datetime.datetime.now().isoformat())
    }
    write_record_to_file('start', start_record)
    # 写入Excel
    excel_record = {
        'write_time': datetime.datetime.now().isoformat(),
        **start_record
    }
    append_to_excel('starts', excel_record)
    # 记录聚合操作
    log_operation('start', session_id, data['employee_id'], step_number=0, name='start', value='', extra='', timestamp=excel_record['timestamp'])
    
    print(f"开始演练: 员工ID={data['employee_id']}, 会话ID={session_id}")
    
    return jsonify({
        'status': 'success',
        'message': '记录成功',
        'session_id': session_id
    })

# API 2: 记录用户进入步骤
@app.route('/api/training/step', methods=['POST'])
def record_step():
    if not verify_api_key():
        return jsonify({'status': 'error', 'message': '无效的API Key'}), 401
    
    is_valid, data = validate_request_data(['session_id', 'employee_id', 'step_number', 'step_name'])
    if not is_valid:
        return jsonify({'status': 'error', 'message': data}), 400
    
    # 检查会话是否存在
    if data['session_id'] not in training_data['sessions']:
        return jsonify({'status': 'error', 'message': '无效的会话ID'}), 400
    
    # 记录步骤信息
    step_record = {
        'session_id': data['session_id'],
        'employee_id': data['employee_id'],
        'step_number': data['step_number'],
        'step_name': data['step_name'],
        'timestamp': data.get('timestamp', datetime.datetime.now().isoformat())
    }
    
    training_data['steps'].append(step_record)
    write_record_to_file('step', step_record)
    append_to_excel('steps', {'write_time': datetime.datetime.now().isoformat(), **step_record})
    # 聚合操作
    log_operation('step', data['session_id'], data['employee_id'], step_number=int(data['step_number']), name=data['step_name'], value='', extra='', timestamp=step_record['timestamp'])
    
    print(f"进入步骤: 员工ID={data['employee_id']}, 步骤={data['step_name']}")
    
    return jsonify({
        'status': 'success',
        'message': '记录成功'
    })

# API 3: 记录用户填写表单数据
@app.route('/api/training/form', methods=['POST'])
def record_form_input():
    if not verify_api_key():
        return jsonify({'status': 'error', 'message': '无效的API Key'}), 401
    
    is_valid, data = validate_request_data(['session_id', 'employee_id', 'step_number', 'field_name', 'field_value'])
    if not is_valid:
        return jsonify({'status': 'error', 'message': data}), 400
    
    # 检查会话是否存在
    if data['session_id'] not in training_data['sessions']:
        return jsonify({'status': 'error', 'message': '无效的会话ID'}), 400
    
    # 记录表单数据
    form_record = {
        'session_id': data['session_id'],
        'employee_id': data['employee_id'],
        'step_number': data['step_number'],
        'field_name': data['field_name'],
        'field_value': data['field_value'],  # 不进行脱敏处理
        'timestamp': data.get('timestamp', datetime.datetime.now().isoformat())
    }
    
    training_data['forms'].append(form_record)
    write_record_to_file('form', form_record)
    append_to_excel('forms', {'write_time': datetime.datetime.now().isoformat(), **form_record})
    # 聚合操作
    log_operation('form', data['session_id'], data['employee_id'], step_number=int(data['step_number']), name=data['field_name'], value=data['field_value'], extra='', timestamp=form_record['timestamp'])
    
    print(f"表单输入: 员工ID={data['employee_id']}, 字段={data['field_name']}, 值={data['field_value']}")
    
    return jsonify({
        'status': 'success',
        'message': '记录成功'
    })

# API 3.1: 上传材料（图片/文件）
@app.route('/api/training/upload', methods=['POST'])
def upload_material():
    if not verify_api_key():
        return jsonify({'status': 'error', 'message': '无效的API Key'}), 401

    # multipart/form-data
    session_id = request.form.get('session_id')
    employee_id = request.form.get('employee_id')
    step_number = request.form.get('step_number')
    field_name = request.form.get('field_name', '')
    file = request.files.get('file')

    if not all([session_id, employee_id, step_number, file]):
        return jsonify({'status': 'error', 'message': '缺少必需字段或文件'}), 400

    if session_id not in training_data['sessions']:
        return jsonify({'status': 'error', 'message': '无效的会话ID'}), 400

    original_filename = file.filename or 'unknown'
    safe_name = secure_filename(original_filename)
    ts = datetime.datetime.now().strftime('%Y%m%dT%H%M%S')
    # 目录 logs/uploads/<employee>/<session>
    target_dir = os.path.join(UPLOADS_DIR, employee_id, session_id)
    os.makedirs(target_dir, exist_ok=True)
    saved_path = os.path.join(target_dir, f"{ts}_{safe_name}")
    file.save(saved_path)

    file_size = os.path.getsize(saved_path)
    mime_type = file.mimetype or ''

    upload_record = {
        'session_id': session_id,
        'employee_id': employee_id,
        'step_number': int(step_number),
        'field_name': field_name,
        'original_filename': original_filename,
        'saved_path': saved_path,
        'file_size_bytes': file_size,
        'mime_type': mime_type,
        'timestamp': datetime.datetime.now().isoformat()
    }

    training_data.setdefault('uploads', []).append(upload_record)
    write_record_to_file('upload', upload_record)
    append_to_excel('uploads', {'write_time': datetime.datetime.now().isoformat(), **upload_record})
    # 聚合操作
    log_operation('upload', session_id, employee_id, step_number=int(step_number), name=field_name, value=original_filename, extra=saved_path, timestamp=upload_record['timestamp'])

    print(f"上传材料: 员工ID={employee_id}, 文件={original_filename}, 保存为={saved_path}")

    return jsonify({'status': 'success', 'message': '上传成功', 'path': saved_path})

# API 4: 记录用户完成演练
@app.route('/api/training/complete', methods=['POST'])
def record_completion():
    if not verify_api_key():
        return jsonify({'status': 'error', 'message': '无效的API Key'}), 401
    
    is_valid, data = validate_request_data(['session_id', 'employee_id', 'verification_code'])
    if not is_valid:
        return jsonify({'status': 'error', 'message': data}), 400
    
    # 检查会话是否存在
    if data['session_id'] not in training_data['sessions']:
        return jsonify({'status': 'error', 'message': '无效的会话ID'}), 400
    
    # 解码验证码（示例性处理）
    try:
        decoded_code = base64.b64decode(data['verification_code']).decode('utf-8')
    except Exception:
        decoded_code = data['verification_code']  # 如果解码失败，使用原始值
    
    # 记录完成信息
    completion_record = {
        'session_id': data['session_id'],
        'employee_id': data['employee_id'],
        'verification_code': decoded_code,
        'timestamp': data.get('timestamp', datetime.datetime.now().isoformat())
    }
    
    training_data['completions'].append(completion_record)
    write_record_to_file('complete', completion_record)
    append_to_excel('completions', {'write_time': datetime.datetime.now().isoformat(), **completion_record})
    # 聚合操作
    log_operation('complete', data['session_id'], data['employee_id'], step_number=5, name='verification_code', value=decoded_code, extra='', timestamp=completion_record['timestamp'])
    
    print(f"完成演练: 员工ID={data['employee_id']}, 验证码={decoded_code}")
    
    return jsonify({
        'status': 'success',
        'message': '演练完成，记录成功'
    })

# API 5: 记录用户关闭弹窗
@app.route('/api/training/close', methods=['POST'])
def record_close():
    if not verify_api_key():
        return jsonify({'status': 'error', 'message': '无效的API Key'}), 401
    
    is_valid, data = validate_request_data(['session_id', 'employee_id', 'step_number'])
    if not is_valid:
        return jsonify({'status': 'error', 'message': data}), 400
    
    # 检查会话是否存在
    if data['session_id'] not in training_data['sessions']:
        return jsonify({'status': 'error', 'message': '无效的会话ID'}), 400
    
    # 记录关闭信息
    close_record = {
        'session_id': data['session_id'],
        'employee_id': data['employee_id'],
        'step_number': data['step_number'],
        'timestamp': data.get('timestamp', datetime.datetime.now().isoformat())
    }
    
    training_data['closures'].append(close_record)
    write_record_to_file('close', close_record)
    append_to_excel('closures', {'write_time': datetime.datetime.now().isoformat(), **close_record})
    # 聚合操作
    log_operation('close', data['session_id'], data['employee_id'], step_number=int(data['step_number']), name='close', value='', extra='', timestamp=close_record['timestamp'])
    
    print(f"关闭弹窗: 员工ID={data['employee_id']}, 步骤={data['step_number']}")
    
    return jsonify({
        'status': 'success',
        'message': '记录成功'
    })

# 用于查看收集到的数据（仅用于测试）
@app.route('/api/training/data', methods=['GET'])
def get_training_data():
    return jsonify(training_data)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)